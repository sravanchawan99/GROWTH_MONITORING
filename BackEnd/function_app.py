import azure.functions as func
import logging
from openpyxl import load_workbook
import numpy as np
import json

app = func.FunctionApp(http_auth_level=func.AuthLevel.ANONYMOUS)

# Load Excel once at startup
wb = load_workbook("Weight_Analysis.xlsx")
sheet = wb.active

# Extract columns into lists
heights, green_weights, red_weights = [], [], []
for row in sheet.iter_rows(min_row=2, values_only=True):  # skip header row
    heights.append(float(row[0]))
    green_weights.append(float(row[1]))
    red_weights.append(float(row[2]))

def get_limits(height: float):
    green = np.interp(height, heights, green_weights)
    red = np.interp(height, heights, red_weights)
    return green, red

def classify(height: float, weight: float) -> str:
    green, red = get_limits(height)
    if weight < red:
        return "Red"
    elif weight > green:
        return "Green"
    else:
        return "Yellow"

@app.route(route="predict_zone", methods=["POST"])
def predict_zone(req: func.HttpRequest) -> func.HttpResponse:
    logging.info("Processing growth zone request...")
    try:
        body = req.get_json()
        height = float(body.get("height"))
        weight = float(body.get("weight"))
    except Exception as e:
        return func.HttpResponse(
            f"Invalid input: {str(e)}",
            status_code=400
        )

    zone = classify(height, weight)
    result = {
        "height": height,
        "weight": weight,
        "zone": zone
    }

    return func.HttpResponse(
        json.dumps(result),
        mimetype="application/json",
        status_code=200
    )
